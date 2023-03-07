import openpyxl

wb = openpyxl.load_workbook('input.xlsx')

ws = wb['MAU']

def gpaScore(math, physics, chemistry):
    return (math + physics + chemistry) / 3;

def sortAToZByName():

    sorted_rows = sorted(ws.rows, key=lambda x: x[2].value) #giá trị họ và tên

    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    for row in sorted_rows:
        ws_new.append([cell.value for cell in row])

    wb_new.save('output.xlsx')

def statistics():
    excellent_count = 0 #giỏi
    good_count = 0 #khá
    average_count = 0 #trung bình
    
    for row in ws.iter_rows(min_row=2):
        average_score = gpaScore(row[4], row[5], row[6]) #giá trị điểm trung bình
        if average_score >= 8.0:
            excellent_count += 1 
        elif average_score >= 6.5:
            good_count += 1 
        else:
            average_count += 1 
    #vị trí dữ liệu thống kê
    ws['Q65'] = excellent_count
    ws['R65'] = good_count
    ws['S65'] = average_count